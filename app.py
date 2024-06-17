from flask import Flask, request, render_template, redirect, url_for
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# Define the path for the Excel file
EXCEL_FILE = 'tickets.xlsx'

# Initialize the Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Employee Name', 'Employee ID', 'Issue', 'Date', 'Time', 'IT Support', 'Resolution', 'Status'])
    workbook.save(EXCEL_FILE)


@app.route('/')
def index():
    return render_template('login.html')


@app.route('/login', methods=['POST'])
def login():
    credential_id = int(request.form['eid'])
    if credential_id == 1001:
        return redirect(url_for('employee'))
    elif credential_id == 2574:
        return redirect(url_for('itsupport'))
    else:
        return 'Invalid credential ID'


@app.route('/employee')
def employee():
    return render_template('employee.html')


@app.route('/submit', methods=['POST'])
def submit():
    # Get form data
    employee_name = request.form['fname']
    employee_id = request.form['eid']
    issue = request.form['eiss']
    date = request.form['edate']
    time = request.form['etime']

    # Load existing workbook and sheet
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    # Append new data
    sheet.append([employee_name, employee_id, issue, date, time, '', '', 'Open'])

    # Save back to Excel
    workbook.save(EXCEL_FILE)

    return redirect(url_for('thankyou'))


@app.route('/delete_ticket', methods=['POST'])
def delete_ticket():
    index = int(request.form['index'])  # This is the row number (1-based index)

    # Load existing workbook and sheet
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    # Check if the index is within the range
    if 1 <= index < sheet.max_row:
        sheet.delete_rows(index + 1)

    # Save back to Excel
    workbook.save(EXCEL_FILE)

    return redirect(url_for('itsupport'))


@app.route('/itsupport')
def itsupport():
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    # Read the data from the sheet
    data = list(sheet.values)
    columns = data[0]
    rows = data[1:]

    # Pass the enumerate function to the template
    return render_template('itsupport.html', columns=columns, rows=rows, enumerate=enumerate)


@app.route('/update_ticket', methods=['POST'])
def update_ticket():
    index = int(request.form['index'])  # This is the row number (1-based index)
    it_support = request.form['it_support']
    resolution = request.form['resolution']
    status = request.form['status']

    # Load existing workbook and sheet
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    # Update ticket (index + 1 because Excel is 1-based, and the first row is headers)
    sheet.cell(row=index + 1, column=6, value=it_support)  # IT Support
    sheet.cell(row=index + 1, column=7, value=resolution)  # Resolution
    sheet.cell(row=index + 1, column=8, value=status)  # Status

    # Save back to Excel
    workbook.save(EXCEL_FILE)

    return redirect(url_for('thankyou'))


@app.route('/thankyou')
def thankyou():
    return render_template('thankyou.html')


if __name__ == '__main__':
    app.run(debug=True)
